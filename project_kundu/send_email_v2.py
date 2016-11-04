#!/usr/bin/env python2.7

#This file sends an email from the gmail account.
#####Phase 1- send the email from one account to the other account
#####Phase 2- Read the input from the user and use that for sending the email including the content.
#####Phase 3- Read the input from the excel sheet, read the name, and send a personalized email to that#email id.
#####Phase 4- Read multiple lines of input from the file and send the email.
####Enhancement- Dont display the password in the raw format while entering at the terminal.
#Phase 5- Read the excel file from the user. Display the error message when gmail doesnt allow to connect due to security settings. Display the message that will be sent to the emails and also ask for confirmation. Also can add a test message email to self before sending it to the list of people.
#Phase 6- You can give an option to the user to either enter emails manually or through the excel sheet.
#Phase 7- Send text messages as well. automatetheboringstuff.com
#Phase 5- Add the GUI to the program.

#Phase 9- Receive email, search emails, delete emails- Automatethestuff.com
#Enhancement- Add the support to send attachments in the email.
#Enhancement- Add different sources from where the data can be read.


""" V2- This program reads a fixed excel file and sends a fixed message to the emails present in the excel sheet """

from smtplib import *
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
import getpass

from openpyxl import *
connect_gmail_status = 0
my_dict= {}
def connect_gmail():
	global mysmtp
	global my_email
	print "Connecting to Gmail server. Please wait"
	mysmtp= SMTP("smtp.gmail.com", 587)
	res= mysmtp.starttls()
	if(res[0] != 220):
		print "Connection failed. Please run again\n"
	res= mysmtp.ehlo()
	if(res[0] != 250):
		print "Connection to the gmail server failed. Please run again\n"
	while 1:
		my_email= raw_input("Enter your email ID: ")
		password = getpass.getpass("\nEnter your password: ")
		
		res= mysmtp.login(my_email, password)
		if(res[0] == 235):
			print "Login successful\n"
			break;
		else:
			print "Login details are wrong. Please enter again\n"
	return res[0]		

def send_email(name, email):
	global connect_gmail_status
	if(connect_gmail_status!=235):
		connect_gmail_status= connect_gmail()
	#rep_email= raw_input("Enter the receipients email ID: ")
	rep_email = email

	msg = MIMEMultipart()
	msg['From'] = "paras"
	msg['To'] = rep_email
	msg['Subject'] = "SUBJECT OF THE MAIL"
 	
	body = "Dear %s, YOUR MESSAGE HERE"% name
	msg.attach(MIMEText(body, 'plain'))


	text = msg.as_string()

	#PARAS- Add the reg expression to check if the email is in proper syntax

	#PARAS- Resolve the subject issue.
	mysmtp.sendmail(my_email, rep_email, text)
	#mysmtp.quit()

def read_excel_sheet():
        global name
        global email

        wb = load_workbook("mywb.xlsx")

        sheet= wb.get_active_sheet()
        max_rows  = sheet.max_row
        max_cols = sheet.max_column
        global my_dict
        for i in range (2, max_rows+1):
                name = sheet.cell(row = i, column= 1).value
                email = sheet.cell(row = i, column= 2).value
                my_dict[name] = email
#                print "name - {0} email {1}".format(sheet.cell(row = i, column= 1).value, sheet.cell(row=i, column=2).value)


		

def read_from_excel_and_mail():
	read_excel_sheet()
        for name, email in my_dict.items():
        	send_email(name, email)
                print "Sending email to %s with email id %s" % (name, email)
	mysmtp.quit()

read_from_excel_and_mail()	

	
