#!/usr/bin/env python2.7
from openpyxl import *

wb = load_workbook("mywb.xlsx")

sheet= wb.get_active_sheet()
max_rows  = sheet.max_row
max_cols = sheet.max_column
my_dict = {}
for i in range (2, max_rows+1):
	name = sheet.cell(row = i, column= 1).value
	email = sheet.cell(row = i, column= 2).value
	my_dict[name] = email
	print "name - {0} email {1}".format(sheet.cell(row = i, column= 1).value, sheet.cell(row=i, column=2).value)


for name, email in my_dict.items():
#	send_email(name, email)
	print name, email
